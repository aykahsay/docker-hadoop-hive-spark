import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def run_financial_analysis():
    print("==================================================")
    print("       MAKE VS. BUY FINANCIAL DECISION MODEL      ")
    print("==================================================")
    
    # 1. Inputs & Assumptions
    tax_rate = 0.30
    discount_rate = 0.10
    machine_cost = 1000000
    salvage_value = 200000
    dep_rate = 0.125
    
    make_costs = [1400000, 1600000, 1800000, 2000000, 2400000]
    buy_costs = [2000000, 2200000, 2400000, 2600000, 3000000]
    
    # 2. Depreciation Schedule (Option A - Make)
    # Reducing Balance (Class IV Machinery - 12.5% rate)
    dep_allowances = []
    book_values = []
    current_bv = machine_cost
    
    for year in range(1, 6):
        dep = current_bv * dep_rate
        dep_allowances.append(dep)
        current_bv = current_bv - dep
        book_values.append(current_bv)
        
    final_book_value = book_values[-1]
    terminal_loss = final_book_value - salvage_value
    terminal_tax_shield = terminal_loss * tax_rate
    
    # 3. Cash Flow Calculations
    # Option A: Make
    make_cash_flows = []
    # Year 0
    make_cash_flows.append(-machine_cost)
    # Years 1-4
    for idx in range(4):
        operating_cost = make_costs[idx]
        tax_shield_cost = operating_cost * tax_rate
        dep_tax_shield = dep_allowances[idx] * tax_rate
        net_cf = -operating_cost + tax_shield_cost + dep_tax_shield
        make_cash_flows.append(net_cf)
    # Year 5 (includes salvage and terminal loss tax shield)
    operating_cost_y5 = make_costs[4]
    tax_shield_cost_y5 = operating_cost_y5 * tax_rate
    dep_tax_shield_y5 = dep_allowances[4] * tax_rate
    net_cf_y5 = -operating_cost_y5 + tax_shield_cost_y5 + dep_tax_shield_y5 + salvage_value + terminal_tax_shield
    make_cash_flows.append(net_cf_y5)
    
    # Option B: Buy
    buy_cash_flows = []
    # Year 0
    buy_cash_flows.append(0)
    # Years 1-5
    for cost in buy_costs:
        net_cf = -cost * (1 - tax_rate)
        buy_cash_flows.append(net_cf)
        
    # 4. NPV Calculations (Present Value of Cash Outflows)
    def calc_npv(cf_list, rate):
        npv_val = cf_list[0] # Year 0 is not discounted
        for idx in range(1, len(cf_list)):
            npv_val += cf_list[idx] / ((1 + rate) ** idx)
        return npv_val
        
    make_npv = calc_npv(make_cash_flows, discount_rate)
    buy_npv = calc_npv(buy_cash_flows, discount_rate)
    net_savings = make_npv - buy_npv
    
    print("\n--- Model Outputs ---")
    print(f"Make Option NPV (PV of Outflows): Shs {make_npv:,.2f}")
    print(f"Buy Option NPV (PV of Outflows):  Shs {buy_npv:,.2f}")
    print(f"Net Savings from Making:          Shs {net_savings:,.2f}")
    print(f"Decision:                         {'MAKE THE COMPONENT' if make_npv > buy_npv else 'BUY THE COMPONENT'}")
    print("==================================================")
    
    # 5. Build Styled Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Analysis"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_family = "Segoe UI"
    font_title = Font(name=font_family, size=16, bold=True, color="0F172A")
    font_section = Font(name=font_family, size=12, bold=True, color="1E293B")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_bold = Font(name=font_family, size=11, bold=True, color="1E293B")
    font_regular = Font(name=font_family, size=11, color="1E293B")
    font_italic = Font(name=font_family, size=9, italic=True, color="64748B")
    
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_accent = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_savings = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # light green
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="1E293B")
    border_thick = Side(border_style="medium", color="1E293B")
    
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    bottom_border_thick = Border(bottom=border_thick)
    total_border = Border(top=border_thin, bottom=border_double)
    
    # Write Title
    ws["B2"] = "MAKE VS. BUY FINANCIAL DECISION MODEL"
    ws["B2"].font = font_title
    
    # Write Assumptions Section
    ws["B4"] = "Key Input Assumptions"
    ws["B4"].font = font_section
    
    assumptions = [
        ("Tax Rate", tax_rate, "0.0%"),
        ("Machine Purchase Cost", machine_cost, "Shs #,##0"),
        ("Machine Salvage Value", salvage_value, "Shs #,##0"),
        ("Depreciation Rate (Reducing Balance)", dep_rate, "0.0%"),
        ("Discount Rate (Cost of Capital)", discount_rate, "0.0%")
    ]
    
    for row_idx, (label, val, fmt) in enumerate(assumptions, 5):
        ws.cell(row=row_idx, column=2, value=label).font = font_regular
        val_cell = ws.cell(row=row_idx, column=3, value=val)
        val_cell.font = font_bold
        val_cell.number_format = fmt
        val_cell.alignment = align_right
        ws.cell(row=row_idx, column=2).border = Border(left=border_thin, top=border_thin, bottom=border_thin)
        val_cell.border = Border(right=border_thin, top=border_thin, bottom=border_thin)
        ws.cell(row=row_idx, column=2).fill = fill_accent
        val_cell.fill = fill_accent
        
    # Option A Table: Make
    ws["B11"] = "Option A: Make the Component (Manufacturing)"
    ws["B11"].font = font_section
    
    headers = ["Line Item", "Year 0", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for col_idx, h in enumerate(headers, 2):
        cell = ws.cell(row=12, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 2 else align_left
        cell.border = Border(top=border_thin, bottom=border_thick)
        
    # Populate Make rows
    make_rows = [
        ("Machine Outlay", "=-C6", 0, 0, 0, 0, 0),
        ("Manufacturing Cost", 0, -make_costs[0], -make_costs[1], -make_costs[2], -make_costs[3], -make_costs[4]),
        ("Tax Shield on Manufacturing Cost", 0, "=-D14*$C$5", "=-E14*$C$5", "=-F14*$C$5", "=-G14*$C$5", "=-H14*$C$5"),
        ("Depreciation Base", 0, "=$C$6", "=D18", "=E18", "=F18", "=G18"),
        ("Depreciation Allowance", 0, "=D16*$C$8", "=E16*$C$8", "=F16*$C$8", "=G16*$C$8", "=H16*$C$8"),
        ("Ending Net Book Value", 0, "=D16-D17", "=E16-E17", "=F16-F17", "=G16-G17", "=H16-H17"),
        ("Tax Shield on Depreciation", 0, "=D17*$C$5", "=E17*$C$5", "=F17*$C$5", "=G17*$C$5", "=H17*$C$5"),
        ("Machine Salvage Value", 0, 0, 0, 0, 0, "=$C$7"),
        ("Terminal Loss Tax Shield", 0, 0, 0, 0, 0, "=MAX(0, H18-H20)*$C$5"),
        ("Net Cash Flow (Make)", "=SUM(C13:C13)", "=SUM(D14:D15)+D19+D20+D21", "=SUM(E14:E15)+E19+E20+E21", "=SUM(F14:F15)+F19+F20+F21", "=SUM(G14:G15)+G19+G20+G21", "=SUM(H14:H15)+H19+H20+H21")
    ]
    
    currency_fmt = "Shs #,##0;[Red](Shs #,##0);\"-\""
    
    for r_idx, row_data in enumerate(make_rows, 13):
        label = row_data[0]
        ws.cell(row=r_idx, column=2, value=label).font = font_bold if r_idx == 22 else font_regular
        for c_idx, val in enumerate(row_data[1:], 3):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_bold if r_idx == 22 else font_regular
            cell.number_format = currency_fmt
            cell.alignment = align_right
            cell.border = box_border
            if r_idx == 22:
                cell.border = total_border
                cell.fill = fill_accent
        if r_idx == 22:
            ws.cell(row=r_idx, column=2).border = total_border
            ws.cell(row=r_idx, column=2).fill = fill_accent
            
    # Option B Table: Buy
    ws["B24"] = "Option B: Buy the Component from Market"
    ws["B24"].font = font_section
    
    for col_idx, h in enumerate(headers, 2):
        cell = ws.cell(row=25, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 2 else align_left
        cell.border = Border(top=border_thin, bottom=border_thick)
        
    buy_rows = [
        ("Purchase Cost (Buy)", 0, -buy_costs[0], -buy_costs[1], -buy_costs[2], -buy_costs[3], -buy_costs[4]),
        ("Tax Shield on Purchase Cost", 0, "=-D26*$C$5", "=-E26*$C$5", "=-F26*$C$5", "=-G26*$C$5", "=-H26*$C$5"),
        ("Net Cash Flow (Buy)", "=SUM(C26:C27)", "=SUM(D26:D27)", "=SUM(E26:E27)", "=SUM(F26:F27)", "=SUM(G26:G27)", "=SUM(H26:H27)")
    ]
    
    for r_idx, row_data in enumerate(buy_rows, 26):
        label = row_data[0]
        ws.cell(row=r_idx, column=2, value=label).font = font_bold if r_idx == 28 else font_regular
        for c_idx, val in enumerate(row_data[1:], 3):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_bold if r_idx == 28 else font_regular
            cell.number_format = currency_fmt
            cell.alignment = align_right
            cell.border = box_border
            if r_idx == 28:
                cell.border = total_border
                cell.fill = fill_accent
        if r_idx == 28:
            ws.cell(row=r_idx, column=2).border = total_border
            ws.cell(row=r_idx, column=2).fill = fill_accent
            
    # Comparison & Decision Section
    ws["B30"] = "Present Value Summary (Discounted Cash Outflows)"
    ws["B30"].font = font_section
    
    decision_rows = [
        ("Present Value of Outflows (Make)", "=NPV($C$9, D22:H22)+C22"),
        ("Present Value of Outflows (Buy)", "=NPV($C$9, D28:H28)+C28"),
        ("Net Savings from Making", "=C31-C32"),
        ("Recommended Action", '=IF(C31>C32, "MAKE THE COMPONENT", "BUY THE COMPONENT")')
    ]
    
    for r_idx, (label, formula) in enumerate(decision_rows, 31):
        ws.cell(row=r_idx, column=2, value=label).font = font_bold if r_idx == 33 or r_idx == 34 else font_regular
        cell_val = ws.cell(row=r_idx, column=3, value=formula)
        cell_val.font = font_bold
        cell_val.alignment = align_right if r_idx != 34 else align_center
        cell_val.border = box_border
        ws.cell(row=r_idx, column=2).border = box_border
        
        if r_idx != 34:
            cell_val.number_format = currency_fmt
        else:
            cell_val.font = Font(name=font_family, size=11, bold=True, color="16A34A")
            
        if r_idx == 33:
            ws.cell(row=r_idx, column=2).fill = fill_savings
            cell_val.fill = fill_savings
        elif r_idx == 34:
            ws.cell(row=r_idx, column=2).fill = fill_savings
            cell_val.fill = fill_savings
            cell_val.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thick)
            ws.cell(row=r_idx, column=2).border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thick)
            
    # Adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'A':
            ws.column_dimensions[col_letter].width = 3
            continue
        for cell in col:
            # Check cell value length
            if cell.value:
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    ws.column_dimensions['B'].width = 42 # Make the label column wide enough
    
    # Save file
    output_dir = "C:\\Users\\Admin\\.gemini\\antigravity\\scratch\\MakeVsBuyDecision"
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, "make_vs_buy_analysis.xlsx")
    wb.save(excel_path)
    print(f"\nSuccessfully saved styled Excel model to: {excel_path}")
    
    # 6. Generate Markdown Report
    report_content = f"""# Make vs. Buy Financial Decision Report

This report evaluates the financial feasibility of making a component in-house using new machinery versus buying it from the market over a 5-year project horizon.

---

## 📊 Summary of Decision Metrics

| Metric | Option A: Make | Option B: Buy | Difference (Savings) |
| :--- | :---: | :---: | :---: |
| **Undiscounted Net Outflows** | Shs {sum(make_cash_flows):,.2f} | Shs {sum(buy_cash_flows):,.2f} | Shs {sum(make_cash_flows) - sum(buy_cash_flows):,.2f} |
| **Present Value of Outflows (10% WACC)** | Shs {make_npv:,.2f} | Shs {buy_npv:,.2f} | **Shs {net_savings:,.2f}** |

> [!IMPORTANT]
> **RECOMMENDED ACTION: MAKE THE COMPONENT**
> In-house manufacturing is the financially superior option. Under a 10% cost of capital (discount rate), making the component results in a Present Value cost of **Shs {make_npv:,.2f}** compared to **Shs {buy_npv:,.2f}** for buying. This represents a net present value saving of **Shs {net_savings:,.2f}**.

---

## 🛠️ Key Assumptions

* **Income Tax Rate**: 30.0% (applied as a tax shield on operating and capital losses/expenses)
* **Discount Rate**: 10.0% WACC
* **Machinery Initial Cost**: Shs 1,000,000 (Class IV wear & tear allowance applied)
* **Machinery Salvage Value (Year 5)**: Shs 200,000
* **Depreciation Rate**: 12.5% reducing balance basis (wear & tear allowance for Class IV machinery)
* **Terminal Loss**: Evaluated at Year 5 based on final book value vs. salvage value.

---

## 📝 Step-by-Step Cash Flow Breakdown

### 1. Option A: Make the Component
* **Year 0**: Machinery capital expenditure of **Shs 1,000,000** (Cash Outflow).
* **Year 1-5 Operating Costs**: Manufacturing costs incurred and reduced by a 30% tax shield.
* **Depreciation Allowance**: Annual tax shield of `Depreciation * 30%` is claimed.
* **Year 5 Terminal Inflow**: Salvage value inflow of **Shs 200,000** plus a tax shield of **Shs {terminal_tax_shield:,.2f}** on the terminal capital loss of **Shs {terminal_loss:,.2f}**.

| Cash Flow Component | Year 0 | Year 1 | Year 2 | Year 3 | Year 4 | Year 5 |
| :--- | :---: | :---: | :---: | :---: | :---: | :---: |
| **Machinery Outlay** | -1,000,000 | 0 | 0 | 0 | 0 | 0 |
| **Manufacturing Cost** | 0 | -1,400,000 | -1,600,000 | -1,800,000 | -2,000,000 | -2,400,000 |
| **Tax Shield on Cost (30%)** | 0 | +420,000 | +480,000 | +540,000 | +600,000 | +720,000 |
| **Depreciation Allowance (12.5%)**| 0 | 125,000 | 109,375 | 95,703 | 83,740 | 73,273 |
| **Tax Shield on Depreciation (30%)**| 0 | +37,500 | +32,813 | +28,711 | +25,122 | +21,982 |
| **Machinery Salvage Value** | 0 | 0 | 0 | 0 | 0 | +200,000 |
| **Terminal Loss Tax Shield** | 0 | 0 | 0 | 0 | 0 | +93,873 |
| **Net Cash Flow (Make)** | **-1,000,000** | **-942,500** | **-1,087,188** | **-1,231,289** | **-1,374,878** | **-1,364,146** |

### 2. Option B: Buy the Component
* Purchasing cost is a tax-deductible expense, resulting in a net cash outflow of `Cost * (1 - Tax Rate) = Cost * 70%`.

| Cash Flow Component | Year 0 | Year 1 | Year 2 | Year 3 | Year 4 | Year 5 |
| :--- | :---: | :---: | :---: | :---: | :---: | :---: |
| **Market Purchase Cost** | 0 | -2,000,000 | -2,200,000 | -2,400,000 | -2,600,000 | -3,000,000 |
| **Tax Shield on Purchase (30%)** | 0 | +600,000 | +660,000 | +720,000 | +780,000 | +900,000 |
| **Net Cash Flow (Buy)** | **0** | **-1,400,000** | **-1,540,000** | **-1,680,000** | **-1,820,000** | **-2,100,000** |

---

## 🔍 Financial Analysis & Interpretation

1. **Operating Cost Advantage**: In-house manufacturing costs are lower than purchase costs every year. Even with Year 0 capital outlay of Shs 1,000,000, the operating savings quickly pay back the machine.
2. **Tax Shields**: The Wear and Tear allowance (Class IV machinery - 12.5% reducing balance) generates tax shields that reduce the net cost of making. At Year 5, selling the machine below book value creates a terminal tax shield of **Shs 93,873**, which further increases Year 5 cash inflows.
3. **Discounted Present Value**: Because the cash outflows for making are lower than buying across all years, the present value of outflows is significantly lower for the Make option. Making the component remains the optimal decision across any reasonable discount rate (WACC).
"""
    
    report_path = os.path.join(output_dir, "make_vs_buy_analysis.md")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_content)
    print(f"Successfully saved detailed markdown report to: {report_path}")
    print("==================================================")

if __name__ == "__main__":
    run_financial_analysis()
